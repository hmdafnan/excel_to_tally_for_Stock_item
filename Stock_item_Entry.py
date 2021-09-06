from tkinter import *
from tkinter import filedialog
import tkinter.font as font
from tkinter import ttk, messagebox
import requests
from openpyxl import load_workbook


class MainWindow:
    #def browseFiles(self):
     #   self.filename = filedialog.askopenfilename(initialdir = "/", title = "Select a File",filetypes = (("Text files", "*.txt*"), ("all files", "*.*")))
      #  opened_filename = Label(root, text = self.filename, width = 100, height = 4, fg = "blue")
       # opened_filename.grid(row=3, column=1)

    def __init__(self, main):
        font_18 = font.Font(family='verdana', size=18)
        font_14 = font.Font(family='verdana', size=14)
        font_10 = font.Font(family='verdana', size=10)

        ##Excel file Path
     #   label_file_explorer = Label(root, text = "Enter Path:", width = 100, height = 4, fg = "blue")
      #  button_explore = Button(root, text = "Browse Files",  command = self.browseFiles)
       # label_file_explorer.grid(row=1, column=1, ipadx=20, ipady=2)
        #button_explore.grid(row=2, column=1, ipadx=0)

        ##Excel file Name
        file_path = Label(main, text="File Name (with format):", font=font_14, fg='green')
        file_path.grid(row=2, column=1, ipady = 10)
        self.file_path_entry = Entry(main, font=font_10, width=20)
        self.file_path_entry.grid(row=2, column=2, ipadx = 20)

        #Stock Item name
        stock_name = Label(main, text="Stock item:", font=font_14, fg="green")
        stock_name.grid(row=4, column=1, ipady=10)
        stock_name_options = ('A','B','C','D','E','F','G','H','I','J','K','L')
        self.stock_name_entry = ttk.Combobox(main, values=stock_name_options, width=5, font=font_18)
        self.stock_name_entry.grid(row=4, column=2)

        #Stock Group name
        stock_group = Label(main, text="Stock Group:", font=font_14, fg="green")
        stock_group.grid(row=5, column=1, ipady=10)
        stock_group_options = ('A','B','C','D','E','F','G','H','I','J','K','L')
        self.stock_group_entry = ttk.Combobox(main, values=stock_group_options, width=5, font=font_18)
        self.stock_group_entry.grid(row=5, column=2)

        #Stock unit
        stock_unit = Label(main, text="Stock Unit:", font=font_14, fg="green")
        stock_unit.grid(row=6, column=1, ipady=10)
        stock_unit_options = ('A','B','C','D','E','F','G','H','I','J','K','L')
        self.stock_unit_entry = ttk.Combobox(main, values=stock_unit_options, width=5, font=font_18)
        self.stock_unit_entry.grid(row=6, column=2)

        #HSN Desc
        hsn = Label(main, text="HSN Desc:", font=font_14, fg="green")
        hsn.grid(row=7, column=1, ipady=10)
        hsn_options = ('A','B','C','D','E','F','G','H','I','J','K','L')
        self.hsn_entry = ttk.Combobox(main, values=hsn_options, width=5, font=font_18)
        self.hsn_entry.grid(row=7, column=2)

        #HSN Code
        hsn_code = Label(main, text="HSN Code:", font=font_14, fg="green")
        hsn_code.grid(row=8, column=1, ipady=10)
        hsn_code_options = ('A','B','C','D','E','F','G','H','I','J','K','L')
        self.hsn_code_entry = ttk.Combobox(main, values=hsn_code_options, width=5, font=font_18)
        self.hsn_code_entry.grid(row=8, column=2)

        #SGST
        sgst = Label(main, text="SGST:", font=font_14, fg="green")
        sgst.grid(row=9, column=1, ipady=10)
        sgst_options = ('A','B','C','D','E','F','G','H','I','J','K','L')
        self.sgst_entry = ttk.Combobox(main, values=sgst_options, width=5, font=font_18)
        self.sgst_entry.grid(row=9, column=2)

        #CGST
        cgst = Label(main, text="CGST:", font=font_14, fg="green")
        cgst.grid(row=10, column=1, ipady=10)
        cgst_options = ('A','B','C','D','E','F','G','H','I','J','K','L')
        self.cgst_entry = ttk.Combobox(main, values=cgst_options, width=5, font=font_18)
        self.cgst_entry.grid(row=10, column=2)

        #IGST
        igst = Label(main, text="IGST:", font=font_14, fg="green")
        igst.grid(row=11, column=1, ipady=10)
        igst_options = ('A','B','C','D','E','F','G','H','I','J','K','L')
        self.igst_entry = ttk.Combobox(main, values=igst_options, width=5, font=font_18)
        self.igst_entry.grid(row=11, column=2)

        ##Row Number
        row_no = Label(main, text="First Row Number:", font=font_14, fg="green")
        row_no.grid(row=12, column=1, ipady=10)
        self.row_no_entry = Entry(main, font=font_18, width=5)
        self.row_no_entry.grid(row=12, column=2)

         # submit button
        submit_button = Button(main, text="Submit", font=font_14, fg='white', bg="green", command=self.Submit)
        submit_button.grid(row=13, column=2, ipadx=20)

    def Submit(self):
        sg = list()
        su = list()
        sg_new = list()
        su_new = list()
        wb = load_workbook(filename=self.file_path_entry.get())
        ws = wb.active
        url = 'http://localhost:9000'

        data = '''
        <ENVELOPE>
            <HEADER>
            <TALLYREQUEST>Import Data</TALLYREQUEST>
            </HEADER>
            <BODY>
            <IMPORTDATA>
            <REQUESTDESC>
                <REPORTNAME>All Masters</REPORTNAME>
            </REQUESTDESC>
            <REQUESTDATA>
        '''

        total = 0
        ##Retrieve and clean-up units
        for stock_unit_iter in ws[self.stock_unit_entry.get()]:
            su.append(stock_unit_iter.value)
            total += 1
        for su_i in su:
            if su_i not in su_new:
                su_new.append(su_i)

        ##Add Stock Units to XML
        for su_iter in su_new:
            data += f'''
            <TALLYMESSAGE xmlns:UDF="TallyUDF">
            <UNIT NAME="{su_iter}" RESERVEDNAME="">
            <NAME>{su_iter}</NAME>
            <ISUPDATINGTARGETID>No</ISUPDATINGTARGETID>
            <ASORIGINAL>Yes</ASORIGINAL>
            <ISGSTEXCLUDED>No</ISGSTEXCLUDED>
            <ISSIMPLEUNIT>Yes</ISSIMPLEUNIT>
            </UNIT>
            </TALLYMESSAGE>
            '''

        ##Retrieve and clean-up the Stock Groups
        for stock_group_iter in ws[self.stock_group_entry.get()]:
            sg.append(stock_group_iter.value)
        for sg_i in sg:
            if sg_i not in sg_new:
                sg_new.append(sg_i)

        ##Add Stock Groups to XML
        for sg_iter in sg_new:
            data += f'''
            <TALLYMESSAGE xmlns:UDF="TallyUDF">
            <STOCKGROUP NAME="{sg_iter}" RESERVEDNAME="">
            <COSTINGMETHOD>Avg. Cost</COSTINGMETHOD>
            <VALUATIONMETHOD>Avg. Price</VALUATIONMETHOD>
            <ISBATCHWISEON>No</ISBATCHWISEON>
            <ISPERISHABLEON>No</ISPERISHABLEON>
            <ISADDABLE>No</ISADDABLE>
            <ISUPDATINGTARGETID>No</ISUPDATINGTARGETID>
            <ASORIGINAL>Yes</ASORIGINAL>
            <IGNOREPHYSICALDIFFERENCE>No</IGNOREPHYSICALDIFFERENCE>
            <IGNORENEGATIVESTOCK>No</IGNORENEGATIVESTOCK>
            <TREATSALESASMANUFACTURED>No</TREATSALESASMANUFACTURED>
            <TREATPURCHASESASCONSUMED>No</TREATPURCHASESASCONSUMED>
            <TREATREJECTSASSCRAP>No</TREATREJECTSASSCRAP>
            <HASMFGDATE>No</HASMFGDATE>
            <ALLOWUSEOFEXPIREDITEMS>No</ALLOWUSEOFEXPIREDITEMS>
            <IGNOREBATCHES>No</IGNOREBATCHES>
            <IGNOREGODOWNS>No</IGNOREGODOWNS>
            <LANGUAGENAME.LIST>
            <NAME.LIST TYPE="String">
                <NAME>{sg_iter}</NAME>
            </NAME.LIST>
            <LANGUAGEID> 1033</LANGUAGEID>
            </LANGUAGENAME.LIST>
            </STOCKGROUP>
            </TALLYMESSAGE>
            '''
        new_cgst = list()
        new_igst = list()
        new_sgst = list()

        cgst2 = 0
        sgst2 = 0
        igst2 = 0
        #Stock Item Master.xlsx

        ##Converting the GST rates to proper format
        for i in range(2, total+1):
            cgst1 = self.cgst_entry.get()
            sgst1 = self.sgst_entry.get()
            igst1 = self.igst_entry.get()
            if ('%' in cgst1) or ('%' in sgst1) or ('%' in igst1):
                if ('%' in cgst1):
                    cgst2 = int(cgst1.strip('%'))
                if ('%' in sgst1):
                    sgst2 = int(sgst1.strip('%'))
                if ('%' in igst1):
                    igst2 = int(igst1.strip('%'))

            elif cgst1 is str:
                cgst2 = 0
                sgst2 = 0
                igst2 = 0

            else:
                cgst2 = cgst1
                sgst2 = sgst1
                igst2 = igst1
            new_cgst.append(cgst2)
            new_sgst.append(sgst2)
            new_igst.append(igst2)


        ##Add Stock Items to XML
        for i in range(2,total+1):
            data += f'''
            <TALLYMESSAGE xmlns:UDF="TallyUDF">
            <STOCKITEM NAME="{ws[self.stock_name_entry.get()+str(i)].value}&#13;&#10;" RESERVEDNAME="">
            <OLDAUDITENTRYIDS.LIST TYPE="Number">
            <OLDAUDITENTRYIDS>1</OLDAUDITENTRYIDS>
            </OLDAUDITENTRYIDS.LIST>
            <PARENT>{ws[self.stock_group_entry.get()+str(i)].value}</PARENT>
            <CATEGORY/>
            <GSTAPPLICABLE>&#4; Applicable</GSTAPPLICABLE>
            <GSTTYPEOFSUPPLY>Goods</GSTTYPEOFSUPPLY>
            <EXCISEAPPLICABILITY>&#4; Applicable</EXCISEAPPLICABILITY>
            <VATAPPLICABLE>&#4; Applicable</VATAPPLICABLE>
            <COSTINGMETHOD>Avg. Cost</COSTINGMETHOD>
            <VALUATIONMETHOD>Avg. Price</VALUATIONMETHOD>
            <BASEUNITS>{ws[self.stock_unit_entry.get()+str(i)].value}</BASEUNITS>
            <VATBASEUNIT>{ws[self.stock_unit_entry.get()+str(i)].value}</VATBASEUNIT>
            <ISCOSTCENTRESON>No</ISCOSTCENTRESON>
            <ISBATCHWISEON>No</ISBATCHWISEON>
            <ISPERISHABLEON>No</ISPERISHABLEON>
            <ISENTRYTAXAPPLICABLE>No</ISENTRYTAXAPPLICABLE>
            <ISCOSTTRACKINGON>No</ISCOSTTRACKINGON>
            <ISUPDATINGTARGETID>No</ISUPDATINGTARGETID>
            <ASORIGINAL>Yes</ASORIGINAL>
            <ISRATEINCLUSIVEVAT>No</ISRATEINCLUSIVEVAT>
            <IGNOREPHYSICALDIFFERENCE>No</IGNOREPHYSICALDIFFERENCE>
            <IGNORENEGATIVESTOCK>No</IGNORENEGATIVESTOCK>
            <TREATSALESASMANUFACTURED>No</TREATSALESASMANUFACTURED>
            <TREATPURCHASESASCONSUMED>No</TREATPURCHASESASCONSUMED>
            <TREATREJECTSASSCRAP>No</TREATREJECTSASSCRAP>
            <HASMFGDATE>No</HASMFGDATE>
            <ALLOWUSEOFEXPIREDITEMS>No</ALLOWUSEOFEXPIREDITEMS>
            <IGNOREBATCHES>No</IGNOREBATCHES>
            <IGNOREGODOWNS>No</IGNOREGODOWNS>
            <CALCONMRP>No</CALCONMRP>
            <EXCLUDEJRNLFORVALUATION>No</EXCLUDEJRNLFORVALUATION>
            <ISMRPINCLOFTAX>No</ISMRPINCLOFTAX>
            <ISADDLTAXEXEMPT>No</ISADDLTAXEXEMPT>
            <ISSUPPLEMENTRYDUTYON>No</ISSUPPLEMENTRYDUTYON>
            <GVATISEXCISEAPPL>No</GVATISEXCISEAPPL>
            <REORDERASHIGHER>No</REORDERASHIGHER>
            <MINORDERASHIGHER>No</MINORDERASHIGHER>
            <ISEXCISECALCULATEONMRP>No</ISEXCISECALCULATEONMRP>
            <INCLUSIVETAX>No</INCLUSIVETAX>
            <GSTCALCSLABONMRP>No</GSTCALCSLABONMRP>
            <MODIFYMRPRATE>No</MODIFYMRPRATE>
            <DENOMINATOR> 1</DENOMINATOR>
            <RATEOFVAT>0</RATEOFVAT>
            <VATBASENO> 1</VATBASENO>
            <VATTRAILNO> 1</VATTRAILNO>
            <VATACTUALRATIO> 1</VATACTUALRATIO>
            <GSTDETAILS.LIST>
            <APPLICABLEFROM>20200401</APPLICABLEFROM>
            <CALCULATIONTYPE>On Value</CALCULATIONTYPE>
            <HSNCODE>{ws[self.hsn_code_entry.get()+str(i)].value}</HSNCODE>
            <HSNMASTERNAME/>
            <HSN>{ws[self.hsn_entry.get()+str(i)].value}&#13;&#10;</HSN>
            <TAXABILITY>Taxable</TAXABILITY>
            <ISREVERSECHARGEAPPLICABLE>No</ISREVERSECHARGEAPPLICABLE>
            <ISNONGSTGOODS>No</ISNONGSTGOODS>
            <GSTINELIGIBLEITC>No</GSTINELIGIBLEITC>
            <INCLUDEEXPFORSLABCALC>No</INCLUDEEXPFORSLABCALC>
            <STATEWISEDETAILS.LIST>
                <STATENAME>&#4; Any</STATENAME>
                <RATEDETAILS.LIST>
                <GSTRATEDUTYHEAD>Central Tax</GSTRATEDUTYHEAD>
                <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                <GSTRATE> {round(float(ws[self.cgst_entry.get()+str(i)].value) * 100)}</GSTRATE>
                </RATEDETAILS.LIST>
                <RATEDETAILS.LIST>
                <GSTRATEDUTYHEAD>State Tax</GSTRATEDUTYHEAD>
                <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                <GSTRATE> {round(float(ws[self.sgst_entry.get()+str(i)].value) * 100)}</GSTRATE>
                </RATEDETAILS.LIST>
                <RATEDETAILS.LIST>
                <GSTRATEDUTYHEAD>Integrated Tax</GSTRATEDUTYHEAD>
                <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                <GSTRATE> {round(float(ws[self.igst_entry.get()+str(i)].value) * 100)}</GSTRATE>
                </RATEDETAILS.LIST>
                <RATEDETAILS.LIST>
                <GSTRATEDUTYHEAD>Cess</GSTRATEDUTYHEAD>
                <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                </RATEDETAILS.LIST>
                <RATEDETAILS.LIST>
                <GSTRATEDUTYHEAD>Cess on Qty</GSTRATEDUTYHEAD>
                <GSTRATEVALUATIONTYPE>Based on Quantity</GSTRATEVALUATIONTYPE>
                </RATEDETAILS.LIST>
                <RATEDETAILS.LIST>
                <GSTRATEDUTYHEAD>State Cess</GSTRATEDUTYHEAD>
                <GSTRATEVALUATIONTYPE>Based on Value</GSTRATEVALUATIONTYPE>
                </RATEDETAILS.LIST>
                <GSTSLABRATES.LIST>        </GSTSLABRATES.LIST>
            </STATEWISEDETAILS.LIST>
            </GSTDETAILS.LIST>
            <LANGUAGENAME.LIST>
            <NAME.LIST TYPE="String">
                <NAME>{ws[self.stock_name_entry.get()+str(i)].value}&#13;&#10;</NAME>
            </NAME.LIST>
            <LANGUAGEID> 1033</LANGUAGEID>
            </LANGUAGENAME.LIST>
            </STOCKITEM>
            </TALLYMESSAGE>
            '''
        data += '''
        </REQUESTDATA>
        </IMPORTDATA>
        </BODY>
        </ENVELOPE>
        '''
        #req = requests.post(url=url, data=data)
        text_file = open('my_xml.xml', 'w')
        n = text_file.write(data)
        text_file.close()

        self.ShowDialog('Your data has been processed.')

        #button_exit = Button(window, text = "Exit", command = exit)

        

    @staticmethod
    def ShowDialog(msg):
        messagebox.showinfo("Information", msg)

root = Tk()
root.columnconfigure(1, minsize=25)
root.rowconfigure(1, minsize=10)
root.title("Excel Stock Item entry")
MainWindow(root)
root.mainloop()